
#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import os
import openpyxl
import uuid
from openpyxl.utils import column_index_from_string


# 获取mac地址，实现本机绑定
addr_num = hex(uuid.getnode())[2:]
mac = "-".join(addr_num[i: i+2] for i in range(0, len(addr_num), 2))
usermac = '90-FB-A6-3F-9C-DF'
if mac != usermac.lower():
    pwd = input('无权限，请输入密码：')
    if pwd != 'ywqh5':
        input('密码不正确，按<回车>退出')
        exit()


while True:
    address = input('输入要对比的文件地址及文件名：')
    fn = os.path.basename(address)

    if fn.endswith('.xls'):
        # （待补充）如果是.xls文件  将起转存为.xlsx 再删掉原xls文件
        pass
    elif not fn.endswith('.xlsx'):
        print('文件格式不正确')
    else:
        break
yn = input('请确定已将要对比的sheet至于前两位，并且参与对比的列位置相同（y/n）')
if yn == 'n':
    print('请将要对比的sheet至于前两位，并保证两个sheet参与对比的列位置相同！')
    input('按 <回车> 退出')
    exit()

columns = input('输入要参与对比的列（无间隔）：')

wb = openpyxl.load_workbook(fn)
sheetone = wb.worksheets[0]
sheettwo = wb.worksheets[1]

# 生成sheet1和sheet2 所有数据的列表 [[第一行数据拼接],[第二行数据拼接],...]
sheetone_data = []
sheettwo_data = []
for i in range(1, sheetone.max_row + 1):
    sheetone_column = []
    for co in columns:
        sheetone_column.append(sheetone.cell(row=i, column=column_index_from_string(co)).value)
    # print(sheetone_column)
    sheetone_data.append(sheetone_column)

for ii in range(1, sheettwo.max_row + 1):
    sheettwo_column = []
    for ct in columns:
        sheettwo_column.append(sheettwo.cell(row=ii, column=column_index_from_string(ct)).value)
    # print(sheettwo_column)
    sheettwo_data.append(sheettwo_column)

# 输出结果到新的sheet中
newsheet = '对比结果'
wb.create_sheet(newsheet)
resultsheet = wb.worksheets[-1]
resultsheet['A1'] = '错误'
resultsheet['B1'] = '行数'
resultsheet['C1'] = '字段内容'

num = 2
for soc in range(len(sheetone_data)):
    if sheetone_data[soc] not in sheettwo_data:
        resultsheet['A' + str(num)] = '在sheet1中不在sheet2中'
        resultsheet['B' + str(num)] = soc + 1
        resultsheet['C' + str(num)] = str(sheetone_data[soc])
        print('表' + wb.sheetnames[0] + '  第' + str(soc + 1) + '行  ' + str(sheetone_data[soc]) + '  不在表' + wb.sheetnames[
            1] + '中')
        num += 1

for stc in range(len(sheettwo_data)):
    if sheettwo_data[stc] not in sheetone_data:
        resultsheet['A' + str(num)] = '在sheet2中不在sheet1中'
        resultsheet['B' + str(num)] = stc + 1
        resultsheet['C' + str(num)] = str(sheettwo_data[stc])
        print('表' + wb.sheetnames[1] + '  第' + str(stc + 1) + '行  ' + str(sheettwo_data[stc]) + '  不在表' + wb.sheetnames[
            0] + '中')
        num += 1

wb.save(filename='（结果）' + fn)
input('按 <回车> 退出')